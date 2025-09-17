import pandas as pd
import openpyxl
from datetime import datetime
import os
import logging
from typing import Dict, List, Optional


class ExcelManager:
    """Gestor para archivos Excel del sistema de captación de viviendas"""
    
    def __init__(self, file_path: str = 'data/viviendas.xlsx'):
        self.file_path = file_path
        self.logger = logging.getLogger(__name__)
        self.columns = [
            'ID',
            'Portal',
            'URL',
            'Titulo',
            'Precio',
            'Ubicacion',
            'Superficie',
            'Habitaciones',
            'Banos',
            'Telefono',
            'Nombre_Contacto',
            'Requiere_Formulario',
            'Fecha_Publicacion',
            'Fecha_Deteccion',
            'Ultima_Actualizacion',
            'Estado',
            'Notas'
        ]
        
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Asegurar que el archivo Excel existe con la estructura correcta"""
        if not os.path.exists(self.file_path):
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
            
            # Crear archivo con estructura inicial
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.file_path, index=False)
            self.logger.info(f"Archivo Excel creado: {self.file_path}")
    
    def load_data(self) -> pd.DataFrame:
        """Cargar datos del archivo Excel"""
        try:
            df = pd.read_excel(self.file_path)
            
            # Asegurar que todas las columnas existen
            for col in self.columns:
                if col not in df.columns:
                    df[col] = None
            
            return df[self.columns]  # Reordenar columnas
            
        except Exception as e:
            self.logger.error(f"Error cargando datos: {e}")
            return pd.DataFrame(columns=self.columns)
    
    def save_data(self, df: pd.DataFrame):
        """Guardar datos al archivo Excel"""
        try:
            # Asegurar que todas las columnas existen
            for col in self.columns:
                if col not in df.columns:
                    df[col] = None
            
            # Reordenar columnas
            df = df[self.columns]
            
            # Guardar con formato
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Viviendas')
                
                # Aplicar formato
                worksheet = writer.sheets['Viviendas']
                self._apply_formatting(worksheet, df)
            
            self.logger.info(f"Datos guardados: {len(df)} registros")
            
        except Exception as e:
            self.logger.error(f"Error guardando datos: {e}")
    
    def _apply_formatting(self, worksheet, df: pd.DataFrame):
        """Aplicar formato al archivo Excel"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Formato de encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num in range(1, len(self.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 8,   # ID
            'B': 12,  # Portal
            'C': 40,  # URL
            'D': 30,  # Titulo
            'E': 12,  # Precio
            'F': 25,  # Ubicacion
            'G': 10,  # Superficie
            'H': 12,  # Habitaciones
            'I': 8,   # Banos
            'J': 15,  # Telefono
            'K': 20,  # Nombre_Contacto
            'L': 15,  # Requiere_Formulario
            'M': 15,  # Fecha_Publicacion
            'N': 15,  # Fecha_Deteccion
            'O': 15,  # Ultima_Actualizacion
            'P': 10,  # Estado
            'Q': 30   # Notas
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
    
    def add_listings(self, new_listings: List[Dict]) -> Dict[str, int]:
        """Añadir nuevos listados al archivo Excel"""
        if not new_listings:
            return {'nuevos': 0, 'actualizados': 0, 'duplicados': 0}
        
        # Cargar datos existentes
        existing_df = self.load_data()
        
        # Convertir nuevos listados a DataFrame
        new_df = pd.DataFrame(new_listings)
        
        stats = {'nuevos': 0, 'actualizados': 0, 'duplicados': 0}
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        for _, row in new_df.iterrows():
            url = row.get('url', '')
            
            # Buscar si ya existe
            existing_match = existing_df[existing_df['URL'] == url]
            
            if existing_match.empty:
                # Nuevo registro
                new_id = self._generate_new_id(existing_df)
                new_record = self._prepare_record(row, new_id, current_time, is_new=True)
                existing_df = pd.concat([existing_df, pd.DataFrame([new_record])], ignore_index=True)
                stats['nuevos'] += 1
                
            else:
                # Actualizar existente
                idx = existing_match.index[0]
                existing_estado = str(existing_df.loc[idx, 'Estado'])  # Obtener estado actual como string
                updated_record = self._prepare_record(row, str(existing_df.loc[idx, 'ID']), current_time, is_new=False, existing_estado=existing_estado)
                
                # Mantener fecha de detección original
                updated_record['Fecha_Deteccion'] = existing_df.loc[idx, 'Fecha_Deteccion']
                
                # Actualizar si hay cambios significativos
                if self._has_significant_changes(existing_df.loc[idx], updated_record):
                    for key, value in updated_record.items():
                        existing_df.loc[idx, key] = value
                    stats['actualizados'] += 1
                else:
                    stats['duplicados'] += 1
        
        # Guardar datos actualizados
        self.save_data(existing_df)
        
        return stats
    
    def _generate_new_id(self, df: pd.DataFrame) -> str:
        """Generar nuevo ID único"""
        if df.empty:
            return "VIV001"
        
        # Obtener el último ID numérico
        existing_ids = df['ID'].dropna().astype(str)
        max_num = 0
        
        for id_str in existing_ids:
            try:
                if id_str.startswith('VIV'):
                    num = int(id_str[3:])
                    max_num = max(max_num, num)
            except ValueError:
                continue
        
        return f"VIV{max_num + 1:03d}"
    
    def _prepare_record(self, row: Dict, record_id: str, current_time: str, is_new: bool, existing_estado: Optional[str] = None) -> Dict:
        """Preparar registro para inserción/actualización"""
        
        # DEBUG: Log del teléfono que llega
        telefono_original = row.get('telefono', '')
        if is_new:  # Solo logear para registros nuevos para evitar spam
            self.logger.info(f"📞 DEBUG ExcelManager: Teléfono recibido = '{telefono_original}' para URL: {row.get('url', 'N/A')}")
        
        # Determinar el estado: preservar el anterior si es Contactado/Descartado, sino Activo
        if is_new or not existing_estado:
            estado = 'Activo'
        elif existing_estado in ['Contactado', 'Descartado']:
            estado = existing_estado  # Preservar estado
            self.logger.info(f"🔒 Preservando estado '{estado}' para inmueble {record_id}")
        else:
            estado = 'Activo'  # Para cualquier otro estado, volver a Activo
        
        record = {
            'ID': record_id,
            'Portal': row.get('portal', ''),
            'URL': row.get('url', ''),
            'Titulo': row.get('titulo', ''),
            'Precio': row.get('precio'),
            'Ubicacion': row.get('ubicacion', ''),
            'Superficie': row.get('superficie', 0),
            'Habitaciones': row.get('habitaciones', 0),
            'Banos': row.get('banos', 0),
            'Telefono': telefono_original,
            'Nombre_Contacto': row.get('nombre_contacto', ''),
            'Requiere_Formulario': 'Sí' if row.get('requiere_formulario', False) else 'No',
            'Fecha_Publicacion': row.get('fecha_publicacion', ''),
            'Fecha_Deteccion': current_time if is_new else None,
            'Ultima_Actualizacion': current_time,
            'Estado': estado,
            'Notas': ''
        }
        
        return record
    
    def _has_significant_changes(self, old_record: pd.Series, new_record: Dict) -> bool:
        """Verificar si hay cambios significativos en el registro"""
        important_fields = ['Precio', 'Telefono', 'Estado']
        
        for field in important_fields:
            if str(old_record.get(field, '')) != str(new_record.get(field, '')):
                return True
        
        return False
    
    def get_statistics(self) -> Dict:
        """Obtener estadísticas de los datos"""
        df = self.load_data()
        
        if df.empty:
            return {}
        
        stats = {
            'total_anuncios': len(df),
            'activos': len(df[df['Estado'] == 'Activo']),
            'con_telefono': len(df[df['Telefono'].notna() & (df['Telefono'] != '')]),
            'requieren_formulario': len(df[df['Requiere_Formulario'] == 'Sí']),
            'por_portal': df['Portal'].value_counts().to_dict(),
            'precio_promedio': df['Precio'].mean() if df['Precio'].notna().any() else 0,
            'ultima_actualizacion': df['Ultima_Actualizacion'].max() if df['Ultima_Actualizacion'].notna().any() else None
        }
        
        return stats
    
    def export_filtered_data(self, filters: Dict, output_path: str) -> bool:
        """Exportar datos filtrados a un nuevo archivo"""
        try:
            df = self.load_data()
            
            # Aplicar filtros
            if filters.get('portal'):
                df = df[df['Portal'] == filters['portal']]
            
            if filters.get('min_price'):
                df = df[df['Precio'] >= filters['min_price']]
            
            if filters.get('max_price'):
                df = df[df['Precio'] <= filters['max_price']]
            
            if filters.get('estado'):
                df = df[df['Estado'] == filters['estado']]
            
            # Exportar
            df.to_excel(output_path, index=False)
            return True
            
        except Exception as e:
            self.logger.error(f"Error exportando datos filtrados: {e}")
            return False
    
    def mark_as_contacted(self, listing_ids: List[str]):
        """Marcar listados como contactados"""
        df = self.load_data()
        
        for listing_id in listing_ids:
            mask = df['ID'] == listing_id
            if mask.any():
                df.loc[mask, 'Estado'] = 'Contactado'
                df.loc[mask, 'Ultima_Actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        self.save_data(df)
    
    def mark_as_discarded(self, listing_ids: List[str]):
        """Marcar listados como descartados"""
        try:
            df = self.load_data()
            
            for listing_id in listing_ids:
                mask = df['ID'] == listing_id
                if mask.any():
                    df.loc[mask, 'Estado'] = 'Descartado'
                    df.loc[mask, 'Ultima_Actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            self.save_data(df)
            self.logger.info(f"Marcados como descartados: {len(listing_ids)} inmuebles")
            return True
            
        except Exception as e:
            self.logger.error(f"Error marcando como descartados: {e}")
            return False
    
    def update_listing_status(self, listing_id: str, new_status: str):
        """Actualizar el estado de un inmueble específico"""
        try:
            df = self.load_data()
            mask = df['ID'] == listing_id
            
            if mask.any():
                df.loc[mask, 'Estado'] = new_status
                df.loc[mask, 'Ultima_Actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.save_data(df)
                self.logger.info(f"Estado actualizado para ID {listing_id}: {new_status}")
                return True
            else:
                self.logger.warning(f"No se encontró inmueble con ID: {listing_id}")
                return False
                
        except Exception as e:
            self.logger.error(f"Error actualizando estado: {e}")
            return False
